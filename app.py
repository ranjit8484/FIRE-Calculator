# app.py
import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from io import BytesIO
from openpyxl.styles import PatternFill
import json
import os

st.set_page_config(layout="wide", page_title="Interactive FIRE Model")

# ---------------------------
# Helpers
# ---------------------------
def pct_to_decimal(val_pct):
    try:
        return float(val_pct) / 100.0
    except:
        return 0.0

def money(x):
    try:
        return "${:,.0f}".format(x)
    except:
        return ""

# Build projection (no buffer spending; only annual_spending + big-ticket)
def build_projection(inputs):
    start_age = int(inputs["starting_age"])
    end_age = int(inputs["projection_end_age"])
    retirement_age = int(inputs["retirement_age"])
    ss_start_age = int(inputs["ss_start_age"])

    ages = list(range(start_age, end_age + 1))
    rows = []

    portfolio = float(inputs["portfolio_at_start"])
    annual_spending_today = float(inputs["annual_spending_today"])
    inflation = float(inputs["inflation"])
    annual_contribution = float(inputs["annual_contribution"])
    pre_ret_return = float(inputs["pre_ret_return"])
    post_ret_return = float(inputs["post_ret_return"])
    ss_annual_benefit = float(inputs["ss_annual_benefit"])
    ss_inflation_adjust = inputs.get("ss_inflation_adjust", True)

    mortgage_balance_remaining = float(inputs["mortgage_balance"])
    mortgage_annual_payment = float(inputs["mortgage_annual_payment"])

    big_expenses = {
        inputs.get("big1_age", 0): inputs.get("big1_amount", 0),
        inputs.get("big2_age", 0): inputs.get("big2_amount", 0),
        inputs.get("big3_age", 0): inputs.get("big3_amount", 0),
    }

    for age in ages:
        row = {}
        row["Age"] = age
        row["Portfolio Start"] = portfolio

        # Contribution until retirement
        contrib = annual_contribution if age < retirement_age else 0.0
        row["Contribution"] = contrib

        # Growth applied to start-of-year portfolio
        ret_rate = pre_ret_return if age < retirement_age else post_ret_return
        growth = portfolio * ret_rate
        row["Growth"] = growth

        # Spending: only after retirement (annual_spending_today inflated from retirement start)
        if age < retirement_age:
            inflated_spend = 0.0
        else:
            years_since_retirement = age - retirement_age
            # Inflate from retirement year (spending is defined in today's dollars)
            inflated_spend = annual_spending_today * ((1 + inflation) ** years_since_retirement)
        row["Inflated Spending"] = inflated_spend

        # Social Security (starts at ss_start_age)
        if age >= ss_start_age:
            if ss_inflation_adjust:
                ss = ss_annual_benefit * ((1 + inflation) ** (age - ss_start_age))
            else:
                ss = ss_annual_benefit
        else:
            ss = 0.0
        row["Social Security"] = ss

        # Mortgage payment: reduce remaining balance by annual mortgage contribution until zero
        if mortgage_balance_remaining > 0:
            mp = min(mortgage_balance_remaining, mortgage_annual_payment)
            mortgage_balance_remaining -= mp
        else:
            mp = 0.0
        row["Mortgage Payment"] = mp

        # Big-ticket one-time expense this year (no inflation applied to the given amount)
        big_ticket = big_expenses.get(age, 0.0)
        row["Big Ticket"] = big_ticket

        # Total spending net of SS plus mortgage and plus big-ticket
        total_spend = inflated_spend - ss + mp + big_ticket
        row["Total Spending"] = total_spend

        # End of year portfolio
        end = portfolio + contrib + growth - total_spend
        row["Portfolio End"] = end

        rows.append(row)
        portfolio = end

    df = pd.DataFrame(rows)
    return df

# Excel export with negative-row highlight
def to_excel_with_highlight(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Projection")
        workbook = writer.book
        worksheet = writer.sheets["Projection"]
        red_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        headers = [cell.value for cell in worksheet[1]]
        try:
            pe_col_idx = headers.index("Portfolio End") + 1
        except ValueError:
            pe_col_idx = None

        for r_idx in range(2, len(df) + 2):
            if pe_col_idx:
                cell_value = worksheet.cell(row=r_idx, column=pe_col_idx).value
                try:
                    if cell_value is not None and float(cell_value) < 0:
                        for c in range(1, len(headers) + 1):
                            worksheet.cell(row=r_idx, column=c).fill = red_fill
                except:
                    pass
    return output.getvalue()

# Styling: convert money columns to $ strings for display and highlight negative Portfolio End rows
def highlight_style(df):
    money_cols = [
        "Portfolio Start","Contribution","Growth",
        "Inflated Spending","Big Ticket",
        "Social Security","Mortgage Payment","Total Spending","Portfolio End"
    ]
    df_formatted = df.copy()
    for col in money_cols:
        if col in df_formatted.columns:
            df_formatted[col] = df_formatted[col].apply(lambda x: "${:,.0f}".format(x))

    def highlight_neg(row):
        try:
            pe = float(row["Portfolio End"].replace('$','').replace(',',''))
        except:
            pe = 0.0
        if pe < 0:
            return ["background-color: #fff2cc" for _ in row]
        else:
            return [""] * len(row)

    sty = df_formatted.style.apply(highlight_neg, axis=1)
    sty = sty.set_properties(**{"text-align": "right"})
    return sty

# ---------------------------
# Sidebar inputs (grouped, clean)
# ---------------------------
st.sidebar.header("Inputs & Assumptions")

with st.sidebar.expander("Personal", expanded=True):
    starting_age = st.number_input("Current age", value=41, min_value=18, max_value=100)
    projection_end_age = st.number_input("Projection end age", value=100, min_value=starting_age, max_value=120)

    # sticky retirement age
    if "retirement_age" not in st.session_state:
        st.session_state.retirement_age = 48
    retirement_age = st.number_input("Retirement age", min_value=starting_age, max_value=projection_end_age,
                                     value=st.session_state.retirement_age)
    st.session_state.retirement_age = retirement_age

    # sticky starting portfolio
    if "portfolio_at_start" not in st.session_state:
        st.session_state.portfolio_at_start = 1550000.0
    portfolio_at_start = st.number_input(f"Portfolio at current age (${starting_age})", value=st.session_state.portfolio_at_start, step=1000.0)
    st.session_state.portfolio_at_start = portfolio_at_start

with st.sidebar.expander("Spending & Contributions", expanded=False):
    annual_spending_today = st.number_input("Annual spending today (today's $)", value=200000, step=1000,
                                           help="This is your target annual spending in today's dollars (will be used after retirement).")
    annual_contribution = st.number_input("Annual contributions pre-retirement ($)", value=100000, step=1000)

with st.sidebar.expander("Returns & Inflation", expanded=False):
    pre_ret_return_pc = st.number_input("Pre-retirement return (%)", value=12.0)
    post_ret_return_pc = st.number_input("Post-retirement return (%)", value=8.0)
    inflation_pc = st.number_input("Inflation (%)", value=3.0)

pre_ret_return = pct_to_decimal(pre_ret_return_pc)
post_ret_return = pct_to_decimal(post_ret_return_pc)
inflation = pct_to_decimal(inflation_pc)

with st.sidebar.expander("Social Security", expanded=False):
    ss_start_age = st.number_input("SS start age", value=67, min_value=62, max_value=75)
    ss_annual_benefit = st.number_input("SS annual benefit today ($)", value=65000, step=1000)
    ss_inflation_adjust = st.checkbox("Inflation adjust Social Security after start age", value=True)

with st.sidebar.expander("Mortgage", expanded=False):
    mortgage_balance = st.number_input("Mortgage balance ($)", value=250000, step=1000)
    mortgage_annual_payment = st.number_input("Annual contribution toward mortgage ($)", value=30000, step=500)

with st.sidebar.expander("Big-ticket Expenses (one-time)", expanded=False):
    big1_age = st.number_input("Big Expense 1 Age", value=55, min_value=starting_age, max_value=projection_end_age)
    big1_amount = st.number_input("Big Expense 1 Amount ($)", value=100000, step=1000)
    big2_age = st.number_input("Big Expense 2 Age", value=57, min_value=starting_age, max_value=projection_end_age)
    big2_amount = st.number_input("Big Expense 2 Amount ($)", value=100000, step=1000)
    big3_age = st.number_input("Big Expense 3 Age", value=80, min_value=starting_age, max_value=projection_end_age)
    big3_amount = st.number_input("Big Expense 3 Amount ($)", value=200000, step=1000)

# -------------------------------------------
# Save / Load / Delete Models (Sidebar)
# -------------------------------------------
st.sidebar.markdown("### Save or Load a Model")

model_name = st.sidebar.text_input("Model Name")

if st.sidebar.button("Save Model"):
    st.session_state.saved_models[model_name] = inputs
    st.sidebar.success("Model saved!")

selected_model = st.sidebar.selectbox("Load Saved Model", [""] + list(st.session_state.saved_models.keys()))

if selected_model != "":
    if st.sidebar.button("Load Model"):
        for key in inputs:
            inputs[key] = st.session_state.saved_models[selected_model][key]
        st.sidebar.success(f"Model {selected_model} loaded!")

if st.sidebar.button("Delete Model"):
    if selected_model in st.session_state.saved_models:
        del st.session_state.saved_models[selected_model]
        st.sidebar.success("Model deleted!")

# ---------------------------
# Build projection and UI outputs
# ---------------------------
df = build_projection(inputs)

# Compute target portfolio at retirement using 4% SWR (user chose 4%)
SWR = 0.04
target_portfolio = annual_spending_today / SWR
# Projected portfolio at retirement (positionally find retirement age row)
age_index = retirement_age - df["Age"].iloc[0]
if 0 <= age_index < len(df):
    projected_at_retirement = df["Portfolio End"].iloc[int(age_index)]
else:
    projected_at_retirement = df["Portfolio End"].iloc[-1]

# Top Goal Dashboard
st.markdown("## Goal Dashboard")
dash_col1, dash_col2, dash_col3 = st.columns([1,1,1])

with dash_col1:
    st.markdown(f"**Age to retire**")
    st.write(f"{retirement_age}")

with dash_col2:
    st.markdown("**Investment needed at retirement (4% SWR)**")
    st.write(money(target_portfolio))

with dash_col3:
    st.markdown("**Portfolio at current age**")
    st.write(money(portfolio_at_start))

dash_col4, dash_col5 = st.columns([1,1])
with dash_col4:
    st.markdown("**Projected portfolio at retirement**")
    st.write(money(projected_at_retirement))
with dash_col5:
    st.markdown("**Portfolio at end age**")
    st.write(money(df["Portfolio End"].iloc[-1]))

st.markdown("---")

# ---------------------------
# Main layout: table (left) and chart/summary (right)
# ---------------------------
col1, col2 = st.columns([2,1])

with col1:
    st.header("Year-by-Year Projection")
    # Table should start with Age column (Age is first column in df)
    styled = highlight_style(df)
    st.dataframe(styled, height=700)
    excel_bytes = to_excel_with_highlight(df)
    st.download_button("Download projection as Excel", excel_bytes, file_name="fire_projection.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

with col2:
    st.header("Chart & Quick Summary")
    ages = df["Age"].to_numpy()
    portfolio_start = df["Portfolio Start"].to_numpy()

    fig, ax = plt.subplots(figsize=(7,5))
    ax.plot(ages, portfolio_start / 1e6, marker="o", label="Portfolio Start (M)", linewidth=2)
    ax.plot(ages, df["Portfolio End"].to_numpy() / 1e6, marker="x", linestyle="--", label="Portfolio End (M)")
    ax.axhline(0, color="red", linestyle=":", label="Zero")
    # highlight negative portfolio_end
    portfolio_end_vals = df["Portfolio End"].to_numpy()
    ax.fill_between(ages, portfolio_end_vals / 1e6, 0, where=(portfolio_end_vals < 0), color="salmon", alpha=0.4)
    ax.set_xlabel("Age")
    ax.set_ylabel("Portfolio (Millions)")
    ax.grid(True, linestyle="--", alpha=0.4)
    ax.legend()
    st.pyplot(fig)

    st.subheader("Quick summary")
    st.markdown(f"**Portfolio at current age ({starting_age})**: {money(portfolio_at_start)}")
    st.markdown(f"**Projected at retirement (age {retirement_age})**: {money(projected_at_retirement)}")
    st.markdown(f"**Target needed (4% SWR)**: {money(target_portfolio)}")
    st.markdown(f"**Portfolio at end age ({projection_end_age})**: {money(df['Portfolio End'].iloc[-1])}")

    negative_rows = df[df["Portfolio End"] < 0]
    if not negative_rows.empty:
        first_neg_age = int(negative_rows.iloc[0]["Age"])
        st.error(f"Portfolio becomes negative at age {first_neg_age}")
    else:
        st.success("Portfolio stays positive through projection end age")

st.markdown("---")
st.write("Model notes")
st.write(
    """
    • Projection starts at current age.
    • Spending starts at retirement and inflates each year using the inflation rate.
    • Big-ticket expenses occur in the specific years you set and are applied as one-time amounts.
    • Social Security can be inflation-adjusted after the chosen start age.
    • Mortgage is automatically reduced each year using your annual contribution until balance reaches 0.
    • Contributions stop at retirement age.
    • Pre- and post-retirement returns are applied to start-of-year portfolio.
    • All monetary values are shown in $ with no decimals.
    """
)
